# -*- coding: utf-8 -*-

"""
  █████████                      ███             █████   
 ███░░░░░███                    ░░░             ░░███    
░███    ░░░   ██████  ████████  ████  ████████  ███████  
░░█████████  ███░░███░░███░░███░░███ ░░███░░███░░░███░   
 ░░░░░░░░███░███ ░░░  ░███ ░░░  ░███  ░███ ░███  ░███    
 ███    ░███░███  ███ ░███      ░███  ░███ ░███  ░███ ███
░░█████████ ░░██████  █████     █████ ░███████   ░░█████ 
 ░░░░░░░░░   ░░░░░░  ░░░░░     ░░░░░  ░███░░░     ░░░░░  
                                      ░███               
                                      █████              
                                     ░░░░░               
Created on Mon Aug 13 13:41:26 2018

@author: 310128142
"""

import sys
import os 
import openpyxl 
import win32com
import time
#sys.setdefaultencoding('utf-8')
PATH =os.path.abspath('.')

def record_path(__path=None, lvl=0 , __file=None):
    c="%s"%hex(0XFFFFFF -(lvl*0x111111 ))[2:]
    
    
    font = openpyxl.styles.Font(size=8, bold=False, name='微软雅黑',  color="1F1F1F")
    
    fill = openpyxl.styles.PatternFill(patternType="solid", start_color=c)
    
    if __path is None:
        __path="E:\010 design folder\036 TCI\TCIdocument"
    if __file is None:
        __file = "dirs.xlsx"
    dir_list=os.path.split(__path)
    level_sign="I"*lvl
    dir_name=dir_list[1]
    log_book=openpyxl.load_workbook(__file)
    sheets=log_book.get_sheet_names()
    log_sheet=log_book.get_sheet_by_name(sheets[0])
    nrow=log_sheet.max_row
    ncol=log_sheet.max_column
    new_list=[level_sign,dir_name]
    log_sheet.append(new_list)
    log_sheet.cell(row=nrow+1,column =2).value = '=HYPERLINK("{}", "{}")'.format(__path, "%s"%dir_name)
    
    log_sheet.cell(row=nrow+1,column =1).font=font
    log_sheet.cell(row=nrow+1,column =1).fill=fill

    log_book.save(__file)
    log_book.close()
    
def record_files(__path,__name,__file): #将信息记录在文件中
    log_book=openpyxl.load_workbook(__file)
    sheets=log_book.get_sheet_names()
    log_sheet=log_book.get_sheet_by_name(sheets[0])
    nrow=log_sheet.max_row
    ncol=log_sheet.max_column
    st_size=os.stat(__path).st_size #获取文件大小
    str_size=size_format(st_size)

    st_mtime=os.stat(__path).st_mtime#获取时间
    str_time =date_format(st_mtime)

    #__path=__path[2:-2]
    new_list=[__name ,__path,str_size,str_time] 
    log_sheet.append(new_list)
    log_sheet.cell(row=nrow+1,column =2).value = '=HYPERLINK("{}", "{}")'.format(__path, "%s"%__path)
    log_book.save(__file)
    log_book.close()
    pass

def size_format(size): #文件size 查询
    if (1024*1024 >size )and (size >1024) :
        return '{:.2f}K'.format(size/1024)
    elif (1024**3 >size )and size >(1024*1024):
        return '{:.2f}M'.format(size/(1024*1024))
    elif(1024**4 >size )and  size >(1024*1024*1024):
        return '{:.2f}G'.format(size/(1024**3))
    else:
        return size

def date_format(date): # 日期格式
    date = time.localtime (date)
    date=time.strftime("%Y-%m-%d %H:%M:%S",date)
    return date
    
def fetch_info(__path=None):
    suffix_list=[".docx",".pdf","doc"]
    
    if __path is None:
        __path="E:\010 design folder\036 TCI\TCIdocument\DHF162180_DNA clock_recovery_considerations.docx"

    file_suffix=os.path.splitext(__path)[1]
    if file_suffix is ".docx":
        word= win32com.client.Dispatch("Word.Application")
        doc=word.Documents.Open(__path)
    print(file_suffix)

def init_record_file(file): #初始化文件
    font = openpyxl.styles.Font(size=10, bold=True, name='微软雅黑',  color="2F4F4F")
    fill = openpyxl.styles.PatternFill(patternType="solid", start_color="33CCFF")
    if file[-5:] != ".xlsx":
        print (file[-5:] )
        raise ValueError("File Name Error!!")

    if os.path.exists(file):
        
        print(file)
        
        os.remove(file)
        init_record_file(file)
    else:
        wb=openpyxl.Workbook()
        ws=wb.active
        ws["A1"].font=font
        ws["A1"].fill=fill
        ws["A1"]="Name"
        ws["B1"].font=font
        ws["B1"].fill=fill
        ws["B1"]="Path"
        ws["C1"].font=font
        ws["C1"].fill=fill
        ws["C1"]="Size"
        ws["D1"].font=font
        ws["D1"].fill=fill
        ws["D1"]="Date"
        
    
   
        
        wb.save(file)
        
def get_file(file,fpath, level=1, mfile_list=None):
    if mfile_list == None:
        mfile_list = []
##列出指定根目录下的所有文件和文件夹
    __parent = os.listdir(fpath)

    for child in __parent:
        child_path = os.path.join(fpath, child)
        
        if os.path.isfile(child_path):
            record_files(child_path,child,file)
        elif os.path.isdir(child_path):
            record_path(child_path,level,file)            
            get_file(file,child_path, level+1)
        else:  
            pass
if __name__ == "__main__":
    init_record_file("dir-list.xlsx")
    get_file("dir-list.xlsx",fpath=PATH)